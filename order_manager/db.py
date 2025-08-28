from __future__ import annotations
import sqlite3
import json
import csv
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

DB_SCHEMA = """
PRAGMA foreign_keys = ON;
CREATE TABLE IF NOT EXISTS clients(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    email TEXT,
    phone TEXT,
    address TEXT
);
CREATE TABLE IF NOT EXISTS products(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    price REAL NOT NULL
);
CREATE TABLE IF NOT EXISTS orders(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    client_id INTEGER NOT NULL REFERENCES clients(id) ON DELETE CASCADE,
    date TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS order_items(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    order_id INTEGER NOT NULL REFERENCES orders(id) ON DELETE CASCADE,
    product_id INTEGER NOT NULL REFERENCES products(id),
    quantity INTEGER NOT NULL
);
"""

class Database:
    def __init__(self, path: str = "orders.sqlite"):
        self.path = path
        self._conn: Optional[sqlite3.Connection] = None

    def connect(self):
        if self._conn is None:
            self._conn = sqlite3.connect(self.path)
            self._conn.row_factory = sqlite3.Row

    @property
    def conn(self) -> sqlite3.Connection:
        self.connect()
        return self._conn

    def init_schema(self):
        self.conn.executescript(DB_SCHEMA)
        self.conn.commit()

    def add_client(self, name: str, email: str, phone: str, address: str):
        row = self.conn.execute(
            "SELECT id, name, email, phone FROM clients WHERE email = ? OR phone = ?",
            (email, phone)
        ).fetchone()
        if row:
            by = []
            if row["email"] == email:
                by.append("email")
            if row["phone"] == phone:
                by.append("телефоном")
            reason = " и ".join(by) if by else "email/телефону"
            raise ValueError(f"Клиент с таким {reason} уже существует (ID={row['id']}, {row['name']}).")

        cur = self.conn.execute(
            "INSERT INTO clients(name, email, phone, address) VALUES (?, ?, ?, ?)",
            (name, email, phone, address)
        )
        self.conn.commit()
        return cur.lastrowid

    def list_clients(self) -> List[sqlite3.Row]:
        return list(self.conn.execute("SELECT * FROM clients ORDER BY id"))

    def find_clients(self, q: str) -> List[sqlite3.Row]:
        q = f"%{q.strip()}%"
        return list(self.conn.execute(
            "SELECT * FROM clients "
            "WHERE name LIKE ? OR email LIKE ? OR phone LIKE ? OR address LIKE ? "
            "ORDER BY id",
            (q, q, q, q)
        ))

    def get_client_by_id(self, cid: int):
        return list(self.conn.execute(
            "SELECT * FROM clients WHERE id = ? ORDER BY id", (int(cid),)
        ))

    def add_product(self, name: str, price: float) -> int:
        cur = self.conn.execute(
            "INSERT INTO products(name,price) VALUES(?,?)", (name, float(price))
        )
        self.conn.commit()
        return cur.lastrowid

    def list_products(self) -> List[sqlite3.Row]:
        return list(self.conn.execute("SELECT * FROM products ORDER BY id"))

    def create_order(self, client_id: int, date: str, items: List[Tuple[int, int]]) -> int:
        cur = self.conn.execute(
            "INSERT INTO orders(client_id,date) VALUES(?,?)", (client_id, date)
        )
        order_id = cur.lastrowid
        self.conn.executemany(
            "INSERT INTO order_items(order_id,product_id,quantity) VALUES(?,?,?)",
            [(order_id, pid, qty) for (pid, qty) in items],
        )
        self.conn.commit()
        return order_id

    def delete_client(self, client_id: int):
        self.conn.execute("DELETE FROM clients WHERE id=?", (client_id,))
        self.conn.commit()

    def delete_order(self, order_id: int):
        self.conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
        self.conn.commit()

    def delete_product(self, product_id: int):
        cnt = self.conn.execute(
            "SELECT COUNT(*) FROM order_items WHERE product_id=?", (product_id,)
        ).fetchone()[0]
        if cnt > 0:
            raise ValueError("Нельзя удалить товар: он используется в заказах")
        self.conn.execute("DELETE FROM products WHERE id=?", (product_id,))
        self.conn.commit()

    def list_orders(self) -> List[sqlite3.Row]:
        sql = """
        SELECT o.id, o.date, c.name AS client,
               SUM(oi.quantity * p.price) AS total,
               SUM(oi.quantity) AS total_qty
        FROM orders o
        JOIN clients c ON c.id = o.client_id
        JOIN order_items oi ON oi.order_id = o.id
        JOIN products p ON p.id = oi.product_id
        GROUP BY o.id, o.date, c.name
        ORDER BY o.date DESC, o.id DESC
        """
        return list(self.conn.execute(sql))

    def order_items(self, order_id: int) -> List[sqlite3.Row]:
        sql = """
        SELECT oi.id, p.name, p.price, oi.quantity, (p.price*oi.quantity) AS line_total
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        WHERE oi.order_id=?
        """
        return list(self.conn.execute(sql, (order_id,)))

    def export_json(self, out_path: str):
        data = {
            "clients": [dict(row) for row in self.list_clients()],
            "products": [dict(row) for row in self.list_products()],
            "orders": [dict(row) for row in self.conn.execute("SELECT * FROM orders")],
            "order_items": [dict(row) for row in self.conn.execute("SELECT * FROM order_items")],
        }
        Path(out_path).write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def import_json(self, in_path: str):
        data = json.loads(Path(in_path).read_text(encoding="utf-8"))
        cur = self.conn
        cur.execute("DELETE FROM order_items")
        cur.execute("DELETE FROM orders")
        cur.execute("DELETE FROM products")
        cur.execute("DELETE FROM clients")
        for t, rows in data.items():
            if t == "clients":
                cur.executemany("INSERT INTO clients(id,name,email,phone,address) VALUES(?,?,?,?,?)",
                                [(r.get("id"), r["name"], r.get("email",""), r.get("phone",""), r.get("address","")) for r in rows])
            elif t == "products":
                cur.executemany("INSERT INTO products(id,name,price) VALUES(?,?,?)",
                                [(r.get("id"), r["name"], float(r["price"])) for r in rows])
            elif t == "orders":
                cur.executemany("INSERT INTO orders(id,client_id,date) VALUES(?,?,?)",
                                [(r.get("id"), r["client_id"], r["date"]) for r in rows])
            elif t == "order_items":
                cur.executemany("INSERT INTO order_items(id,order_id,product_id,quantity) VALUES(?,?,?,?)",
                                [(r.get("id"), r["order_id"], r["product_id"], r["quantity"]) for r in rows])
        self.conn.commit()

    def export_csv(self, folder: str = "export_csv", *, encoding: str = "cp1251", excel_friendly: bool = True):
        p = Path(folder);
        p.mkdir(exist_ok=True)

        def dump(sql: str, fname: str):
            rows = list(self.conn.execute(sql))
            with (p / fname).open("w", newline="", encoding=encoding, errors="replace") as f:
                if excel_friendly:
                    f.write("sep=;\n")
                w = csv.writer(f, delimiter=';' if excel_friendly else ',', quoting=csv.QUOTE_MINIMAL)
                if rows:
                    headers = list(rows[0].keys())
                    w.writerow(headers)
                    for r in rows:
                        out = []
                        for h in headers:
                            val = r[h]
                            if excel_friendly and h in ("phone", "date") and val is not None:
                                val = f"'{val}"
                            out.append(val)
                        w.writerow(out)
                else:
                    w.writerow([])

        dump("SELECT * FROM clients", "clients.csv")
        dump("SELECT * FROM products", "products.csv")
        dump("SELECT * FROM orders", "orders.csv")
        dump("SELECT * FROM order_items", "order_items.csv")



    def export_xlsx(self, out_path: str = "export.xlsx"):
        wb = Workbook();
        wb.remove(wb.active)

        def add_sheet(title: str, sql: str, formats=None, widths=None):
            ws = wb.create_sheet(title)
            rows = list(self.conn.execute(sql))
            if not rows:
                return
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append([r[h] for h in headers])

            formats = formats or {}
            for col_name, fmt in formats.items():
                if col_name in headers:
                    c = headers.index(col_name) + 1
                    for cell in ws.iter_rows(min_row=2, min_col=c, max_col=c):
                        for x in cell: x.number_format = fmt

            widths = widths or {}
            for col_name, w in widths.items():
                if col_name in headers:
                    c = headers.index(col_name) + 1
                    ws.column_dimensions[get_column_letter(c)].width = w

        add_sheet("clients", "SELECT * FROM clients",
                  widths={"name": 20, "email": 28, "phone": 16, "address": 22})
        add_sheet("products", "SELECT * FROM products",
                  formats={"price": "0.00"}, widths={"name": 24, "price": 12})
        add_sheet("orders", "SELECT * FROM orders",
                  formats={"date": "yyyy-mm-dd"}, widths={"date": 14})
        add_sheet("order_items", "SELECT * FROM order_items")

        wb.save(out_path)

    def seed_demo(self):
        if self.list_clients() or self.list_products():
            return
        a = self.add_client("Иван", "ivan@example.com", "+79991112233", "Москва")
        b = self.add_client("Пётр", "petr@example.com", "+79995556677", "Санкт-Петербург")
        c = self.add_client("Артем", "art@example.com", "+79998887766", "Москва")
        m = self.add_product("Мыло", 100)
        s = self.add_product("Сосиски", 200)
        w = self.add_product("Вода", 55)
        self.create_order(a, "2025-08-13", [(m, 1)])
        self.create_order(a, "2025-08-13", [(s, 1)])
        self.create_order(b, "2025-08-13", [(m, 1), (s, 1)])
        self.create_order(c, "2025-08-14", [(w, 3)])
